import os
import sys
import openpyxl
import csv

csv_path = sys.argv[1] + os.sep + 'FilteredCSV' + os.sep
info_path = sys.argv[1] + os.sep + 'lang.info'
spreadsheet_path = sys.argv[2]
silent = sys.argv[3] == 'on'
dry_run = sys.argv[4] == 'on'

if silent:
   sys.stdout = open(os.devnull, 'w')

files = [file for file in os.listdir(csv_path) if os.path.isfile(os.path.join(csv_path, file))]

workbook = openpyxl.Workbook()

header_fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor='004D80')
header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
alignment_center = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrapText=True)
alignment_wrap = openpyxl.styles.Alignment(vertical='center', wrapText=True)

# Get language list from lang.info
lang_list = []
with open(info_path) as info_file:
    lang_list = info_file.readline().split(',')

# Sort by alphabet
files.sort()

for file in files:
    # Example of file name '01 - Table Name-Foo', Use 2nd index when split by '-'
    number, sheet_name, _ = file.split('-')
    sheet_name = number + '-' + sheet_name
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.page_setup.fitToWidth = 1

    with open(csv_path + file) as csvfile:
        reader = csv.reader(csvfile, delimiter=',', quotechar='"')
        
        # skip header
        next(reader)

        # cell property
        worksheet.column_dimensions['A'].hidden = True
        worksheet.column_dimensions['B'].width = 50
        worksheet.column_dimensions['C'].width = 50

        # create header
        worksheet['A1'].value = 'Key'
        worksheet['A1'].font = header_font
        worksheet['A1'].alignment = alignment_center
        worksheet['A1'].fill = header_fill
        worksheet['B1'].value = 'Original Text'
        worksheet['B1'].font = header_font
        worksheet['B1'].alignment = alignment_center
        worksheet['B1'].fill = header_fill

        # Create header
        col_index = 3
        for lang in lang_list:
            # Skip base language
            if lang == 'base':
                continue

            cell = worksheet.cell(row=1, column=col_index)
            cell.value = 'Translated Text - ' + lang.split('_')[-1].upper() 
            cell.font = header_font
            cell.alignment = alignment_center
            cell.fill = header_fill
            col_index += 1

        # Create values
        row_index = 2
        for datas in reader:
            col_index = 1
            # Apply base text

            for data in datas:
                cell = worksheet.cell(row=row_index, column=col_index)
                cell.value = data
                cell.alignment = alignment_wrap
                col_index += 1
            row_index += 1

# Remove dummy sheet, create by initial
workbook.remove(workbook['Sheet'])
workbook.save(spreadsheet_path)
print('file ' + spreadsheet_path + ' created.')